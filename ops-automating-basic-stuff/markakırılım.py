import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font


def process_excel_file(input_file_path, output_file_path):


    print(" Dosya okunuyor...")
    df = pd.read_excel(input_file_path, sheet_name='counts', header=None)

    categories = []
    for col in range(len(df.columns)):
        if pd.notna(df.iloc[0, col]) and df.iloc[0, col] != '':
            categories.append({
                'name': df.iloc[0, col],
                'start_col': col
            })

    print(f" Bulunan kategoriler: {[cat['name'] for cat in categories]}")

    for i in range(len(categories)):
        if i < len(categories) - 1:
            categories[i]['end_col'] = categories[i + 1]['start_col']
        else:
            categories[i]['end_col'] = len(df.columns)

    all_channels = set()
    for category in categories:
        for col in range(category['start_col'], category['end_col'], 2):
            if col < len(df.columns) and pd.notna(df.iloc[1, col]) and df.iloc[1, col] != '':
                all_channels.add(df.iloc[1, col])

    print(f"🏢 Bulunan markalar: {sorted(list(all_channels))}")

    channels_data = {channel: [] for channel in all_channels}

    for category in categories:
        print(f"🔄 {category['name']} kategorisi işleniyor...")

        category_channels = []
        for col in range(category['start_col'], category['end_col'], 2):
            if col < len(df.columns) and pd.notna(df.iloc[1, col]) and df.iloc[1, col] != '':
                category_channels.append({
                    'name': df.iloc[1, col],
                    'col': col
                })

        for channel in category_channels:
            for row in range(2, len(df)):
                marka = df.iloc[row, channel['col']]
                count = df.iloc[row, channel['col'] + 1] if channel['col'] + 1 < len(df.columns) else None

                if pd.notna(marka) and marka != '' and pd.notna(count) and count != '':
                    channels_data[channel['name']].append({
                        'Kanal': channel['name'],
                        'Kategori': category['name'],
                        'Marka': marka,
                        'Haziran Adet': count
                    })

    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:

        for channel_name in sorted(all_channels):
            print(f" {channel_name} sheet'i oluşturuluyor...")

            if channels_data[channel_name]:
                # DataFrame oluştur ve sheet'e yaz
                channel_df = pd.DataFrame(channels_data[channel_name])
                channel_df.to_excel(writer, sheet_name=channel_name, index=False)
                print(f"   {len(channels_data[channel_name])} satır oluşturuldu")
            else:
                empty_df = pd.DataFrame(columns=['Kanal', 'Kategori', 'Marka', 'Haziran Adet'])
                empty_df.to_excel(writer, sheet_name=channel_name, index=False)
                print(f"   Veri bulunamadı - boş sheet oluşturuldu")

    print("🎨 Formatlar uygulanıyor...")
    workbook = openpyxl.load_workbook(output_file_path)

    red_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Başlık satırını formatla (1. satır)
        for col in range(1, 5):  # A1:D1
            cell = sheet.cell(row=1, column=col)
            cell.fill = red_fill
            cell.font = white_font

        # Sütun genişliklerini ayarla
        sheet.column_dimensions['A'].width = 15  # Kanal
        sheet.column_dimensions['B'].width = 15  # Kategori
        sheet.column_dimensions['C'].width = 25  # Marka
        sheet.column_dimensions['D'].width = 12  # Haziran Adet

    workbook.save(output_file_path)
    print(f"💾 Dosya kaydedildi: {output_file_path}")


def main():
    """
    """

    input_file = "kırılım2.xlsx"  # Giriş dosyası
    output_file = "markalara_gore_kirilim2.xlsx"  # Çıkış dosyası

    try:
        process_excel_file(input_file, output_file)
        print("\n🎉 İşlem tamamlandı!")
        print(f"📁 Çıkış dosyası: {output_file}")
        print("🏢 Her marka için ayrı sheet oluşturuldu:")
        print("   - Beymen")
        print("   - Boyner")
        print("   - FashFed")
        print("   - Vakko")
        print("   - Vakkorama")
        print("🎨 Başlık formatları uygulandı (kırmızı arka plan)")

        print("\ Her sheet'te format:")
        print("   Kanal | Kategori | Marka | Haziran Adet")
        print("\ Örnek:")
        print("   Beymen | Aksesuar | 40 Million Eyewear | 59")
        print("   Beymen | Ayakkabı | Academia | 73")
        print("   Beymen | Çanta | AllSaints | 46")

    except Exception as e:
        print(f" Hata oluştu: {str(e)}")


if __name__ == "__main__":
    main()

# Kullanım:
# 1. Bu dosyayı .py uzantısıyla kaydet (örn: marka_processor.py)
# 2. Terminal/cmd'de: pip install pandas openpyxl
# 3. Terminal/cmd'de: python marka_processor.py
# 4. markalara_gore_kirilim.xlsx dosyası oluşacak

# Sonuç:
#  5 sheet: Beymen, Boyner, FashFed, Vakko, Vakkorama
#  Her sheet'te o markaya ait TÜM kategorilerden veriler
#  Format: Kanal | Kategori | Marka | Haziran Adet
#  Kırmızı başlık formatı